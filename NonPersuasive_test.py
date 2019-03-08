import csv
from persuasiveTool import textCheck
import numpy as np
import xlrd
import openpyxl

sheetname = "Sheet"
modelName = ["LogisticRegression","LinearDiscriminantAnalysis","GaussianNB","SupportVectorMachine"]
# data = []
# result=[]
workbook = xlrd.open_workbook("Nonpersuasive TFIDF Analysis.xlsx", on_demand=True)
sheet = workbook.sheet_by_name(sheetname)
arrayofvalues = sheet.col_values(3,1)

wb =openpyxl.load_workbook(filename = 'Nonpersuasive TFIDF Analysis.xlsx')
sheet_ranges = wb[sheetname]


countColNum = 8
rowNUm = 2
flag = 0
print("start print")
#iteration of the col
for item in arrayofvalues:
    #print(item)
    count = 0
    modelColNum = 9
    #go through 4 models
    for modelResult in textCheck(item, modelName):
        #print the result of each model
        print(modelResult)
        #write
        sheet_ranges.cell(row=rowNUm,column=modelColNum).value=modelResult[0]
        modelColNum += 1
        #count the first element in the result
        if modelResult[0] == 1.0:
            count += 1

     #if statement to determine the output
        print("count:{}".format(count))

    if count >= 3:
        persuasiveResult = 2
        flag += 1
    elif count == 2:
        persuasiveResult = 1
        flag += 1
    else:
        persuasiveResult = 0
     # print("count:")
     # print(count)
    print("pause,check persuasiveResult")
    print(persuasiveResult)
    sheet_ranges.cell(row=rowNUm, column=countColNum).value = persuasiveResult
    rowNUm += 1
    print("#####################################")
     # csv_write = csv.writer(out, dialect="excel")
     # temp = [item, rowresult]
     # csv_write.writerow(temp)
wb.save("Nonpersuasive TFIDF Analysis.xlsx")
# you save to a wrong file
# the previous file has been modified
# you need yo re-run it
# and generate a new persuasive xlsx
# also the non persuasive
print("flag:{}".format(flag))